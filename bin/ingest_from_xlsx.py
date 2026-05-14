"""bin/ingest_from_xlsx.py
multisite-platform 原生 · 投手 FB xlsx 上传 → 自动 ingest

输入:  requests/uploads/{投手}-*.xlsx
输出:  data/uploads/{投手}-*.json(结构化中间产物,审计用)
       触发 数据分析 agent 跑该投手视角

xlsx 期望结构(参考 crave-AI ingest_from_xlsx.py 反推):
  · 第 1 行表头
  · 含 'Ad set name' 列(广告组名,内含"广告组NN")
  · 含 'Amount spent' 列(花费 USD)
  · 可选含 HVU 列(若 FB 后台有此自定义指标)

文件名约定:
  HZM-2026-05-14.xlsx          → owner=HZM, date=2026-05-14
  HZM-FB-export-2026-05-14.xlsx → owner=HZM, date=2026-05-14
  CHJ-2026-05-14.xlsx          → owner=CHJ
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
UPLOADS_DIR = REPO / "requests" / "uploads"
PROCESSED_DIR = REPO / "data" / "uploads"

DEFAULT_SITE = "elysianu"
DEFAULT_CHANNEL = "FB"
KNOWN_CHANNELS = {"FB", "Google", "Twitter", "TikTok"}


# ─── 文件名 → owner/site/channel/date ────────────────
# 新格式(三维): {owner}-{site}-{channel}-{date}-{原名}.xlsx
#   例: HZM-elysianu-FB-2026-05-14-原名.xlsx
# 兼容旧格式: {owner}-{date}.xlsx / {owner}-FB-{date}.xlsx 等
FILENAME_RE_FULL = re.compile(
    r"^(?P<owner>HZM|CHJ|HNN|ZXR|LZL|PLZ)-"
    r"(?P<site>[a-z0-9_-]+)-"
    r"(?P<channel>FB|Google|Twitter|TikTok)-"
    r"(?P<date>\d{4}-\d{2}-\d{2})",
    re.IGNORECASE,
)
FILENAME_RE_LEGACY = re.compile(
    r"^(?P<owner>HZM|CHJ|HNN|ZXR|LZL|PLZ)"
    r".*?(?P<date>\d{4}-\d{2}-\d{2})?",
    re.IGNORECASE,
)


def parse_filename(name: str) -> tuple[str, str, str, str]:
    """从文件名抽 owner / site / channel / date。
    新格式优先,旧格式回退默认 site=elysianu / channel=FB。
    """
    stem = Path(name).stem
    m = FILENAME_RE_FULL.match(stem)
    if m:
        # 渠道做大小写归一(枚举严格)
        ch = m.group("channel")
        for k in KNOWN_CHANNELS:
            if ch.lower() == k.lower():
                ch = k
                break
        return m.group("owner").upper(), m.group("site").lower(), ch, m.group("date")

    # 旧格式回退
    m = FILENAME_RE_LEGACY.match(stem)
    today = datetime.now().strftime("%Y-%m-%d")
    if not m:
        return "unknown", DEFAULT_SITE, DEFAULT_CHANNEL, today
    return (
        (m.group("owner") or "unknown").upper(),
        DEFAULT_SITE,
        DEFAULT_CHANNEL,
        m.group("date") or today,
    )


# ─── xlsx 解析 ─────────────────────────
def parse_xlsx(path: Path) -> dict:
    """
    返回:
      {
        "rows": [{"group_id": "组17", "spend": 50.95, "hvu": 47}, ...],
        "header_meta": {"adset_col": "Ad set name", "spend_col": "Amount spent", ...},
        "warnings": [...]
      }
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl 未安装,跑 pip3 install openpyxl"}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # 列索引识别(支持中英文 + 部分泰文)
    adset_col = spend_col = hvu_col = None
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_lower = str(h).lower()
        if "ad set name" in h_lower or "广告组名" in str(h):
            if adset_col is None:
                adset_col = i
        if "amount spent" in h_lower or "花费" in str(h) or "已使用金额" in str(h):
            if spend_col is None:
                spend_col = i
        if h_lower in ("hvu", "high value users", "高价值用户"):
            hvu_col = i

    warnings = []
    if adset_col is None:
        return {"error": f"找不到 'Ad set name' 列 · 表头={headers}"}
    if spend_col is None:
        return {"error": f"找不到 'Amount spent' 列 · 表头={headers}"}

    # 聚合(同一组多行 → 总和)
    group_data: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[adset_col] is None:
            continue
        m = re.search(r"广告组(\d+)|组(\d+)|group\s*(\d+)", str(row[adset_col]), re.IGNORECASE)
        if not m:
            warnings.append(f"行无组号: {str(row[adset_col])[:50]}")
            continue
        gid = f"组{m.group(1) or m.group(2) or m.group(3)}"
        spend = float(row[spend_col]) if row[spend_col] else 0
        hvu = int(row[hvu_col]) if hvu_col is not None and row[hvu_col] else 0

        if gid not in group_data:
            group_data[gid] = {"group_id": gid, "spend": 0, "hvu": 0}
        group_data[gid]["spend"] += spend
        group_data[gid]["hvu"] += hvu

    # 计算 cphq
    rows = []
    for g in group_data.values():
        cphq = round(g["spend"] / g["hvu"], 2) if g["hvu"] > 0 else None
        g["cphq"] = cphq
        rows.append(g)

    return {
        "rows": rows,
        "header_meta": {
            "adset_col": headers[adset_col],
            "spend_col": headers[spend_col],
            "hvu_col": headers[hvu_col] if hvu_col is not None else None,
        },
        "warnings": warnings,
    }


# ─── 主入口 ────────────────────────────
def process_xlsx(path: Path) -> dict:
    """处理一个 xlsx · 返回处理结果"""
    owner, site, channel, date = parse_filename(path.name)
    parsed = parse_xlsx(path)

    if "error" in parsed:
        return {"ok": False, "owner": owner, "site": site, "channel": channel,
                "date": date, "error": parsed["error"]}

    # 落 JSON(审计 + 给 agent 用)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    out_path = PROCESSED_DIR / f"{owner}-{site}-{channel}-{date}-{datetime.now().strftime('%H%M%S')}.json"
    out_path.write_text(json.dumps({
        "_meta": {
            "owner": owner, "site": site, "channel": channel, "date": date,
            "source_file": path.name,
            "ingested_at": datetime.now().isoformat(),
            "header_meta": parsed["header_meta"],
        },
        "rows": [{**r, "owner": owner, "site": site, "channel": channel, "date": date}
                 for r in parsed["rows"]],
        "warnings": parsed["warnings"],
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    # 触发数据分析 agent
    sys.path.insert(0, str(REPO / "platform-core/执行层"))
    sys.path.insert(0, str(REPO / "platform-core/中间件"))
    sys.path.insert(0, str(REPO / "platform-core/数据层"))
    sys.path.insert(0, str(REPO / "platform-core/引擎层"))

    import importlib.util
    p = REPO / "platform-core/执行层/数据分析/agent.py"
    spec = importlib.util.spec_from_file_location("agent_数据分析", p)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    from _base import AgentInput, TimeWindow
    from engine import get_engine

    engine = get_engine()
    daily_rows = [
        {"group": r["group_id"], "group_id": r["group_id"], "owner": owner,
         "site": site, "channel": channel,
         "date": date, "spend": r["spend"], "hvu": r["hvu"], "cphq": r["cphq"]}
        for r in parsed["rows"]
    ]

    upstream = {
        "daily_rows": daily_rows, "owner_filter": owner,
        "site_filter": site, "channel_filter": channel,
        "data_date": date, "uploaded_xlsx": path.name,
    }
    tid = engine.submit("数据分析", site, priority=2, lookback_days=1,
                       upstream_output=upstream, event_id=f"xlsx_upload_{path.stem}")
    engine.next()
    inp = AgentInput(site_id=site, time_window=TimeWindow.last_n_days(1),
                     overrides={"owner_filter": owner, "channel_filter": channel},
                     upstream_output=upstream)
    result = mod.run(inp)
    engine.complete(tid, result)

    return {
        "ok": True, "owner": owner, "site": site, "channel": channel, "date": date,
        "rows_count": len(parsed["rows"]),
        "agent_status": result.status.value,
        "json_output": str(out_path.relative_to(REPO)),
        "warnings": parsed["warnings"],
    }


def scan_uploads() -> list[dict]:
    """扫 requests/uploads/ 下所有 xlsx,逐个 ingest"""
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    results = []
    for f in sorted(UPLOADS_DIR.glob("*.xlsx")):
        print(f"\n→ 处理 {f.name}")
        r = process_xlsx(f)
        results.append({"file": f.name, **r})
        if r["ok"]:
            # 归档:移到 _done/
            done = UPLOADS_DIR / "_done"
            done.mkdir(exist_ok=True)
            f.rename(done / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{f.name}")
            print(f"  ✓ {r['rows_count']} 组 · agent {r['agent_status']} · 归档")
        else:
            print(f"  ✗ {r['error']}")
    return results


if __name__ == "__main__":
    if len(sys.argv) > 1:
        # 单文件模式
        r = process_xlsx(Path(sys.argv[1]))
        print(json.dumps(r, ensure_ascii=False, indent=2))
    else:
        # 扫 uploads/ 全部
        rs = scan_uploads()
        print(f"\n本轮处理 {len(rs)} 个 xlsx")
