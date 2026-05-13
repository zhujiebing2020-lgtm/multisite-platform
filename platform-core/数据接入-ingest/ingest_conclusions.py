"""
Ingest conclusions from xlsx dashboard: classify landing pages, check thresholds, output pending conclusions.

Usage:
    python3 renderer/ingest_conclusions.py --dash ~/Desktop/广告组数据看板0507xlsx.xlsx
"""
from __future__ import annotations
import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

OUTPUT_PATH = Path.home() / "Documents/GitHub/crave-AI/data/conclusions_pending.json"

THRESHOLDS = {
    "文章详情页": 15,
    "剧情体验页": 15,
    "商品单品页": 8,
    "商品总览页": 5,
}

ADVERTISER_THRESHOLD = 6


def classify_landing_page(lp: str) -> str:
    lp = lp.strip()
    if "scene" in lp or "exp-app" in lp:
        return "剧情体验页"
    elif "ai-sex-toys" in lp:
        parts = lp.split("ai-sex-toys")
        after = parts[1] if len(parts) > 1 else ""
        if after and after not in ["", "/", "…"]:
            return "商品单品页"
        else:
            return "商品总览页"
    elif "multisensory" in lp or re.match(r'\d{3,4}-', lp):
        return "文章详情页"
    else:
        return "文章详情页"


def parse_xlsx(dash_path: Path):
    wb = openpyxl.load_workbook(dash_path, data_only=True)
    ws = wb["广告组数据看板"]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    date_cols = {}
    for i, h in enumerate(headers):
        if h and re.match(r'\d+/\d+', str(h)):
            date_cols[str(h)] = i

    groups = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        adv = ws.cell(row=row, column=2).value
        lp = ws.cell(row=row, column=3).value
        if not name or "合计" in str(name):
            continue

        gm = re.match(r'(组\d+)', str(name))
        if not gm:
            continue

        lp_str = str(lp or "")
        lp_type = classify_landing_page(lp_str)

        total_spend = 0
        total_hvu = 0
        active_days = 0
        for d, col_idx in date_cols.items():
            cell = ws.cell(row=row, column=col_idx + 1).value
            if cell:
                m = re.match(r'\$?([\d.]+)/([\d]+)/', str(cell))
                if m:
                    total_spend += float(m.group(1))
                    total_hvu += int(m.group(2))
                    active_days += 1

        if total_spend == 0:
            continue

        lifetime_cphq = total_spend / total_hvu if total_hvu > 0 else None
        is_green = lifetime_cphq is not None and lifetime_cphq <= 3 and total_spend >= 30

        groups.append({
            "group": gm.group(1),
            "advertiser": adv,
            "landing_type": lp_type,
            "total_spend": total_spend,
            "total_hvu": total_hvu,
            "lifetime_cphq": lifetime_cphq,
            "active_days": active_days,
            "is_green": is_green,
        })

    return groups


def compute_conclusions(groups):
    by_type = defaultdict(list)
    for g in groups:
        by_type[g["landing_type"]].append(g)

    conclusions = []

    # 落地页类型对比
    type_data = {}
    for lp_type, threshold in THRESHOLDS.items():
        type_groups = by_type.get(lp_type, [])
        qualified = [g for g in type_groups if g["total_spend"] >= 30]
        n = len(qualified)
        reached = n >= threshold

        if qualified:
            total_spend = sum(g["total_spend"] for g in qualified)
            total_hvu = sum(g["total_hvu"] for g in qualified)
            cphq = total_spend / total_hvu if total_hvu > 0 else None
            green_count = sum(1 for g in qualified if g["is_green"])
            green_rate = green_count / n if n > 0 else 0

            type_data[lp_type] = {
                "n": n,
                "threshold": threshold,
                "reached": reached,
                "total_spend": round(total_spend, 2),
                "total_hvu": total_hvu,
                "cphq": round(cphq, 2) if cphq else None,
                "green_count": green_count,
                "green_rate": round(green_rate * 100, 1),
            }
        else:
            type_data[lp_type] = {"n": 0, "threshold": threshold, "reached": False}

    all_reached = all(type_data.get(t, {}).get("reached", False) for t in ["文章详情页", "剧情体验页"])
    conclusions.append({
        "type": "落地页类型对比",
        "status": "达到门槛" if all_reached else "未达门槛",
        "data": type_data,
        "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
    })

    # 投手差异
    by_adv = defaultdict(list)
    for g in groups:
        if g["landing_type"] == "文章详情页" and g["total_spend"] >= 30:
            by_adv[g["advertiser"]].append(g)

    adv_data = {}
    for adv, adv_groups in by_adv.items():
        n = len(adv_groups)
        total_spend = sum(g["total_spend"] for g in adv_groups)
        total_hvu = sum(g["total_hvu"] for g in adv_groups)
        cphq = total_spend / total_hvu if total_hvu > 0 else None
        green_count = sum(1 for g in adv_groups if g["is_green"])
        adv_data[adv] = {
            "n": n,
            "cphq": round(cphq, 2) if cphq else None,
            "green_count": green_count,
            "green_rate": round(green_count / n * 100, 1) if n > 0 else 0,
        }

    adv_reached = any(d["n"] >= ADVERTISER_THRESHOLD for d in adv_data.values())
    conclusions.append({
        "type": "投手操作能力差异",
        "status": "达到门槛" if adv_reached else "未达门槛",
        "data": adv_data,
        "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
    })

    return conclusions


def print_status(conclusions):
    print("[结论检查]", end=" ")
    for c in conclusions:
        if c["type"] == "落地页类型对比":
            for lp_type, d in c["data"].items():
                n = d.get("n", 0)
                threshold = d.get("threshold", 0)
                if d.get("reached"):
                    print(f"{lp_type}: {n}组 ✅", end="  ")
                elif n > 0:
                    diff = threshold - n
                    print(f"{lp_type}: {n}组 ⏳(差{diff}组)", end="  ")
                else:
                    print(f"{lp_type}: 0组 ❌", end="  ")
    print()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dash", required=True, help="Path to xlsx dashboard file")
    args = parser.parse_args()

    dash_path = Path(args.dash)
    if not dash_path.exists():
        print(f"ERROR: {dash_path} not found")
        sys.exit(1)

    groups = parse_xlsx(dash_path)
    conclusions = compute_conclusions(groups)

    output = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "xlsx_source": dash_path.name,
        "conclusions": conclusions,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"Output: {OUTPUT_PATH}")
    print_status(conclusions)


if __name__ == "__main__":
    main()
