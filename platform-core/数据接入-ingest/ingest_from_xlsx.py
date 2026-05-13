"""Ingest ad spend data from advertiser Excel files + HVU JSON into the metric runtime."""
from __future__ import annotations
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))

from core.metric_graph.dag import MetricDAG
from core.propagation.engine import PropagationEngine
from core.state_store.sqlite import SQLiteStateStore


SCHEMAS_DIR = Path(__file__).parent.parent / "schemas"
DATA_BASE = Path("/Users/zhujiebing/Library/Mobile Documents/com~apple~CloudDocs/crave ai/3_执行层-人工维护/单平台分析/Facebook/原始数据-投手Excel与CSV")


ADVERTISER_MAP = {
    "陈洪骏": "CHJ",
    "黄梓铭": "HZM",
    "黄楠楠": "HNN",
}


def extract_group_number(adset_name: str) -> str | None:
    m = re.search(r"广告组(\d+)", adset_name)
    if m:
        return f"组{m.group(1)}"
    return None


def ingest_spend_file(engine: PropagationEngine, filepath: Path, advertiser: str, date: str):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    adset_col = None
    spend_col = None
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_lower = str(h).lower()
        if "ad set name" in h_lower or "ชื่อชุดโฆษณา" in h_lower:
            if adset_col is None:
                adset_col = i
        if "amount spent" in h_lower or "จำนวนเงินที่ใช้จ่ายไป" in h_lower:
            spend_col = i

    if adset_col is None or spend_col is None:
        print(f"  WARNING: Could not find columns in {filepath.name}")
        return 0

    group_spend = defaultdict(float)
    for row in ws.iter_rows(min_row=2, values_only=True):
        adset = row[adset_col]
        spend = row[spend_col]
        if adset and spend:
            group = extract_group_number(str(adset))
            if group:
                group_spend[group] += float(spend)

    count = 0
    for group, total_spend in group_spend.items():
        engine.handle_event({
            "event_type": "ad_spend_report",
            "amount": round(total_spend, 2),
            "ad_group": group,
            "advertiser": advertiser,
            "date": date,
        })
        count += 1

    return count


def ingest_hvu_json(engine: PropagationEngine, filepath: Path, date: str):
    with open(filepath) as f:
        data = json.load(f)

    sessions = data.get("sessions", [])
    group_hvu = defaultdict(lambda: defaultdict(int))

    for s in sessions:
        traffic = s.get("traffic", {})
        utm_source = str(traffic.get("utm_source", "")).lower()
        if utm_source != "facebook":
            continue

        utm_campaign = traffic.get("utm_campaign", "")
        utm_content = traffic.get("utm_content", "")

        advertiser = None
        if "chj" in utm_campaign.lower() or utm_content.startswith("C0"):
            advertiser = "CHJ"
        elif "hnn" in utm_campaign.lower() or utm_content.startswith("Hn"):
            advertiser = "HNN"
        elif "hzm" in utm_campaign.lower() or utm_content.startswith("Mm"):
            advertiser = "HZM"

        if not advertiser:
            continue

        group_hvu[advertiser]["__total__"] += 1

    count = 0
    for adv, groups in group_hvu.items():
        total = groups["__total__"]
        for _ in range(total):
            engine.handle_event({
                "event_type": "hvu_session",
                "session_id": f"hvu_{adv}_{count}",
                "ad_group": f"__all_{adv}__",
                "advertiser": adv,
                "date": date,
            })
            count += 1

    return count


def main():
    date = sys.argv[1] if len(sys.argv) > 1 else "2026-05-06"
    db_path = str(Path(__file__).parent.parent / "crave_metrics.db")

    print(f"Ingesting data for {date}")
    print(f"DB: {db_path}")
    print()

    dag = MetricDAG.load_from_yaml(SCHEMAS_DIR / "metrics.yaml")
    store = SQLiteStateStore(db_path)
    engine = PropagationEngine(dag, store)

    spend_files = list(DATA_BASE.glob(f"*5.6.xlsx"))
    print(f"Found {len(spend_files)} spend files")

    total_groups = 0
    for f in spend_files:
        for cn_name, code in ADVERTISER_MAP.items():
            if cn_name in f.name:
                print(f"  Processing {f.name} → {code}")
                n = ingest_spend_file(engine, f, code, date)
                total_groups += n
                print(f"    {n} groups ingested")
                break

    hvu_file = DATA_BASE / f"FB—HVU—0506.json"
    if hvu_file.exists():
        print(f"\n  Processing HVU: {hvu_file.name}")
        n = ingest_hvu_json(engine, hvu_file, date)
        print(f"    {n} HVU sessions ingested")

    print(f"\nDone. Total groups: {total_groups}, DB version: {store.global_version}")
    store.close()


if __name__ == "__main__":
    main()
